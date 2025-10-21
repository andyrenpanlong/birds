from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
from bs4 import BeautifulSoup
from urllib.parse import quote
import base64
import requests
import csv
import os

headers = {
    'authority': 'www.google.com.hk',
    'method': 'GET',
    # 'path': '/search?sca_esv=280a44b9c274f648&q=Afghan+Snowfinch&udm=2&fbs=AEQNm0AaBOazvTRM_Uafu9eNJJzCjPEAP5HX2BE31zy5nlFpWvns3PnbKLxkXEhI71m1WxuyGXhNUouYPTHnpIch5xJgIMcZof5GjjJE9giO3Q868M8bOV1cSsJN7w0Zrg98RFguyHOZm3xFt2oGVWl-pJTEpoK9LGg_7usrFKfWPNc2HJFGR3WY8McCLmP6UHbReplFnP4m&sa=X&ved=2ahUKEwje3qSzn-WJAxXwZfUHHc5oIEAQtKgLegQIEhAB&biw=1364&bih=753&dpr=2',
    'scheme': 'https',
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'accept-encoding': 'gzip, deflate, br, zstd',
    'accept-language': 'zh-CN,zh;q=0.9',
    'cache-control': 'max-age=0',
    # 'cookie': 'AEC=AZ6Zc-VomWV9U-iwzrmKaW7jtU0b_qPE_vKIHed1I9oqA5X9T24hq-5kug; NID=519=bg5487oEiLe0XbM1WKh-sTXD_sYZCB8SjGkfJJ_sNenCT_oymcsYYegdVCp17dR-KkkrED2cADjIeSRRKciLUbUIvla5lRDLXZBcnvRzpzNuci4zEknnPuJDXT0YUx0B8wvgl_Ao2V1qSA-nEg40sXSSrxCfv2sc2Go4yWNOOPTlp7lkO93ZRGj2gsN6WD8OSSU6QIEQgwKUCmXBdguKLWd_0Z6dJkCa8So; DV=A5cuHIutxIwcMMNiN6YW-gb0bxXfMxk',
    'priority': 'u=0, i',
    'referer': 'https://www.google.com.hk/',
    'sec-ch-prefers-color-scheme': 'dark',
    'sec-ch-ua': '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
    'sec-ch-ua-arch': '"x86"',
    'sec-ch-ua-bitness': '"64"',
    'sec-ch-ua-form-factors': '"Desktop"',
    # 'sec-ch-ua-full-version': '"130.0.6723.117"',
    'sec-ch-ua-full-version-list': '"Chromium";v="130.0.6723.117", "Google Chrome";v="130.0.6723.117", "Not?A_Brand";v="99.0.0.0"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-model': '""',
    'sec-ch-ua-platform': '"macOS"',
    'sec-ch-ua-platform-version': '"13.7.1"',
    'sec-ch-ua-wow64': '?0',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'same-origin',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36'
}

num = 0
# 读取文档数据，提取关键词
def read_xlsx():
    # 加载工作簿
    workbook = load_workbook(filename='birds_count.xlsx')
    # 获取默认的活动工作表
    worksheet = workbook.active
    keywords = []
    key_val_all = []
    # 读取单元格的值
    for row in worksheet.iter_rows(values_only=True):
        keywords.append(row[0])
        key_val_all.append(row)
    return keywords

def get_google_img(keyword):
    url = 'https://www.google.com.hk/search?sca_esv=280a44b9c274f648&q={}&udm=2&fbs=AEQNm0AaBOazvTRM_Uafu9eNJJzCjPEAP5HX2BE31zy5nlFpWvns3PnbKLxkXEhI71m1WxuyGXhNUouYPTHnpIch5xJgIMcZof5GjjJE9giO3Q868M8bOV1cSsJN7w0Zrg98RFguyHOZm3xFt2oGVWl-pJTEpoK9LGg_7usrFKfWPNc2HJFGR3WY8McCLmP6UHbReplFnP4m&sa=X&ved=2ahUKEwje3qSzn-WJAxXwZfUHHc5oIEAQtKgLegQIEhAB&biw=1364&bih=753&dpr=2'.format(keyword)
    response = requests.get(url, headers=headers)
    bs = BeautifulSoup(response.text, 'html5lib')
    # 获取top3的图片
    top3_ids = []
    top3_imgs = []
    top3_imgs2 = []
    print('132131:', len(bs.select('#search g-img')))
    for img in bs.select('#search g-img'):
        img_id = img.select('img')[0].get('id')
        # if len(img.select('img')[0].get('width')) >= 3:
        top3_ids.append(img_id)
    # 获取完整的base64 img
    for pic_id in top3_ids:
        for i in bs.select('script'):
            if 'data:image' in str(i) and "var ii=['{}']".format(pic_id) in str(i):
                base64_img = str(i).split('var s=')[1].split("';var")[0].replace("'", '')
                top3_imgs2.append(base64_img)
                if base64_img not in top3_imgs:
                    top3_imgs.append(base64_img)
    print('999999999:', len(top3_imgs), len(top3_imgs2))
    return top3_imgs

def save_comments_csv(data_arr):
    with open('comments_new_6.csv', 'a+', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerow(data_arr)  # 向文件中追加字符串列表
    print('insert ok！！！')

# 创建空目录
def create_path(your_path, base64_image_lists):
    # 假设固定目录为 images 文件夹，位于当前工作目录下
    fixed_directory = './google图片/{}'.format(your_path)
    # 如果目录不存在则创建
    if not os.path.exists(fixed_directory):
        os.makedirs(fixed_directory)
    # Base64 编码的图片数据（这里只是示例数据，实际应用中你需要从其他地方获取这个数据）
    for base64_image_data in base64_image_lists:
        base64_image_data = base64_image_data.replace('\\x3d', '=')
        # 提取图像格式
        image_format = base64_image_data.split(';')[0].split('/')[1]
        # 解码 Base64 数据
        image_data = base64.b64decode(base64_image_data.split(',')[1])
        global num
        num += 1
        # 生成唯一的文件名，这里简单使用当前时间戳
        import time
        timestamp = str(int(time.time()))
        file_name = f"{timestamp}_{num}.{image_format}"
        # 完整的文件路径
        file_path = os.path.join(fixed_directory, file_name)
        # 保存图片
        with open(file_path, "wb") as f:
            f.write(image_data)
        print(f"图片已保存到 {file_path}")

data_lists = read_xlsx()
all_lens = len(data_lists)
# 序号从1开始，不算表头
# for i in range(1, all_lens):
for i in range(1, 2):
    keyword1 = str(data_lists[i])
    print('232432:', keyword1)
    keyword = quote(keyword1.replace('&', ''))
    base64_image_lists = get_google_img(keyword)
    # print(base64_image_lists)
    # save_comments_csv([i, keyword1, base64_image_lists])
    create_path(keyword1, base64_image_lists)

