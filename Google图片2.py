from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
from bs4 import BeautifulSoup
from urllib.parse import quote
import base64
import requests
import csv
import os

headers = {
    'authority': 'www.google.com.hk',
    'method': 'GET',
    # 'path': '/search?q=Binske+-+9D4+-+Live+Resin+Disposable,+350mg&newwindow=1&sca_esv=ee61b556728f00f2&sca_upv=1&hl=zh-CN&source=hp&biw=1707&bih=940&ei=rT7hZsXXGa690-kPxf71gAU&iflsig=AL9hbdgAAAAAZuFMvWeygEklQQ77_AS2SAua8XGr5BmB&ved=0ahUKEwiF4oa1p7qIAxWu3jQHHUV_HVAQ4dUDCAc&uact=5&oq=Binske+-+9D4+-+Live+Resin+Disposable,+350mg&gs_lp=EgNpbWciK0JpbnNrZSAtIDlENCAtIExpdmUgUmVzaW4gRGlzcG9zYWJsZSwgMzUwbWdIiglQywFYywFwAHgAkAEAmAFGoAFGqgEBMbgBA8gBAPgBAvgBAYoCC2d3cy13aXotaW1nmAIAoAIAqAIAmAMBkgcAoAcz&sclient=img&udm=2',
    'scheme': 'https',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Cache-Control': 'max-age=0',
    # 'Cookie': 'NID=517=zjjqMpHKAGXiXPe7XTfpiUkvpYiYWwEa_ezbeNxxfmPXUSSDVMfiiBq8hKYCMjQ4LiljBFb8Mees_bfLVoZxTALP8f3cLpqvLn8UZxSUPJOV70d9V0-v0bFgaddA-jDE1z-R7BoE07r_YMKM3lZd08CQiqKqb_kF8yGBTwyKYdUGQZzQkp-rhqD2FTzAC82btQK4xplk_Rtivg',
    'Priority': 'u=0, i',
    'Referer': 'https://images.google.com.hk/',
    'Sec-Ch-Prefers-Color-Scheme': 'light',
    'Sec-Ch-Ua': '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
    'Sec-Ch-Ua-Arch': '"x86"',
    'Sec-Ch-Ua-Bitness': '"64"',
    'Sec-Ch-Ua-Full-Version': '"124.0.6367.92"',
    'Sec-Ch-Ua-Full-Version-List': '"Chromium";v="124.0.6367.92", "Google Chrome";v="124.0.6367.92", "Not-A.Brand";v="99.0.0.0"',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Ch-Ua-Model': '""',
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Sec-Ch-Ua-Platform-Version': '"10.0.0"',
    'Sec-Ch-Ua-Wow64': '?0',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-site',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
}

num = 0
# 读取文档数据，提取关键词
def read_xlsx():
    # 加载工作簿
    workbook = load_workbook(filename='birds_count.xlsx')
    # 获取默认的活动工作表
    worksheet = workbook.active
    keywords = []
    key_val_all = []
    # 读取单元格的值
    for row in worksheet.iter_rows(values_only=True):
        keywords.append(row[0])
        key_val_all.append(row)
    return keywords

url = 'https://www.google.com.hk/search?vet=12ahUKEwiut7Ol4eWJAxWJ7DQHHdbeESg4FBDErwIoAHoECAEQAw..i&ved=2ahUKEwiI2eTK4OWJAxUor1YBHaJMB14Qqq4CegQIChAE&bl=Kgt7&s=images&opi=89978449&sca_esv=339a018dea4bc811&udm=2&yv=3&q=Anambra+Waxbill&newwindow=1&biw=1707&bih=879&sxsrf=ADLYWIIep3IrLfojxQ20IRhqi6b2eYl_rw:1731928539087&ei=2yE7Z8j_BKje2roPopmd8AU&start=30&sa=N&asearch=arc&cs=0&async=arc_id:srp_2yE7Z8j_BKje2roPopmd8AU_130,ffilt:all,ve_name:MoreResultsContainer,use_ac:false,_id:arc-srp_2yE7Z8j_BKje2roPopmd8AU_130,_pms:s,_fmt:pc,_basejs:%2Fxjs%2F_%2Fjs%2Fk%3Dxjs.s.zh.ZEDG-AvAx4s.2018.O%2Fam%3DAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACAAAAAAAAAAAAACKCAgAAAAAAAgAIQAAAAAAAAAAAAAAAAAAAAQAICEAgJAAAQAAAAAAAsAACAwAACBAAAAgAAAAAIEAAAAAgQAez3HwcAAAAAAAAAAABCACACAAAAAC4AAAgADQEAIAAIAAAADAAAAAgAIAAAACgAABBAAAAAACAAAAAAAAAAAAAAAAQA9AMAAAAAAAAAAAAAAAQAAAAAAGgABQAC-AEACAAAAAAAAAAAICAAAMARMAABEAAAAAAA4D4AeDwgHFJYAAAAAAAAAAAAAAAAQAASBHNA-gsCQAAAAAAAAAAAAAAAAAAAQIqgicsNAAo%2Fdg%3D0%2Fbr%3D1%2Frs%3DACT90oFn79oFuQTWy3z6_FHyhfzchT2n7g,_basecss:%2Fxjs%2F_%2Fss%2Fk%3Dxjs.s.1RouNDlYloc.L.B1.O%2Fam%3DAKQOAEIAAAAIAACIABUABAAAAAAAAAAAAAAAAAAAAAAAAEgAAACAEAAAAAAAIAAAACIAAAAEAAAQvAAQwC4AQAAAwAcAgKMCwAAAAAIAEAAJAAAAAAAAAgAoEAAAgAAAAAAgIAAAIAAAEQCAAAAKAQAAIAYGAAAwAAAAAQBBAAEACGAAFCAACQgACAH0owIAAAQADAAAIAgIJxgGECqAMOAgAAAAAAAAAAAAAEAAgBAAAASgAAAIIABADwABYAAA0AxEAIAgwEAARQAQAAAAAAAAAAAQAIEAAEDABMARMAABAAAAAAAAABIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAoAAAAAAAAAAAAAAAAAAAAAAAg%2Fbr%3D1%2Frs%3DACT90oHfnb-KzVr1I5MuNui3tBfqQ-5yPA,_basecomb:%2Fxjs%2F_%2Fjs%2Fk%3Dxjs.s.zh.ZEDG-AvAx4s.2018.O%2Fck%3Dxjs.s.1RouNDlYloc.L.B1.O%2Fam%3DAKQOAEIAAAAIAACIABUABAAAAAAAAAAAAAAAAAAAAAAAAEgAAACAEAAAAAAAIAAACKKAgAAEAAAQvAIQwC4AQAAAwAcAgKMCwAAAQAICEAgJAAAQAAAAAgAsEACAwAACBAAgIgAAIAAIEQCAAAgaAez3PwcGAAAwAAAAAQBDACECCGAAFC4ACQgADQH0owIIAAQADAAAIAgIJxgGECqAMPBgAAAAACAAAAAAAEAAgBAAAASg9AMIIABADwABYAAA0AxEAIAgwGgARQAS-AEACAAAAAAQAIEAIGDABMARMAABEAAAAAAA4D4AeDwgHFJYAAAAAAAAAAAAAAAAQAASBHNA-gsCQAAAAAAAAAAAAAAAAAAAQIqgicsNAAo%2Fd%3D1%2Fed%3D1%2Fdg%3D0%2Fbr%3D1%2Fujg%3D1%2Frs%3DACT90oGy8hrBykHbgzPn2HktrSfUNp1pzw'
url = 'https://www.google.com.hk/search?vet=12ahUKEwja0t3O8OWJAxUXiq8BHWN9DvMQxK8CegQIFBAC..i&ved=2ahUKEwja0t3O8OWJAxUXiq8BHWN9DvMQqq4CegQIFBAE&bl=Mhqh&s=images&opi=89978449&sca_esv=339a018dea4bc811&udm=2&yv=3&q=Anambra+Waxbill&newwindow=1&biw=1707&bih=879&sxsrf=ADLYWIJdEDSA9kP1OS5Wz49rpUvgHTeFaQ:1731932842329&ei=qjI7Z9rVE5eUvr0P4_q5mA8&start=10&sa=N&asearch=arc&cs=0&async=arc_id:srp_qjI7Z9rVE5eUvr0P4_q5mA8_110,ffilt:all,ve_name:MoreResultsContainer,use_ac:false,_id:arc-srp_qjI7Z9rVE5eUvr0P4_q5mA8_110,_pms:s,_fmt:pc,_basejs:%2Fxjs%2F_%2Fjs%2Fk%3Dxjs.s.zh.ZEDG-AvAx4s.2018.O%2Fam%3DAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACAAAAAAAAAAAAACKCAgAAAAAAAgAIQAAAAAAAAAAAAAAAAAAAAQAICEAgJAAAQAAAAAAAsAACAwAACBAAAAgAAAAAIEAAAAAgQAez3HwcAAAAAAAAAAABCACACAAAAAC4AAAgADQEAIAAIAAAADAAAAAgAIAAAACgAABBAAAAAACAAAAAAAAAAAAAAAAQA9AMAAAAAAAAAAAAAAAQAAAAAAGgABQAC-AEACAAAAAAAAAAAICAAAMARMAABEAAAAAAA4D4AeDwgHFJYAAAAAAAAAAAAAAAAQAASBHNA-gsCQAAAAAAAAAAAAAAAAAAAQIqgicsNAAo%2Fdg%3D0%2Fbr%3D1%2Frs%3DACT90oFn79oFuQTWy3z6_FHyhfzchT2n7g,_basecss:%2Fxjs%2F_%2Fss%2Fk%3Dxjs.s.1RouNDlYloc.L.B1.O%2Fam%3DAKQOAEIAAAAIAACIABUABAAAAAAAAAAAAAAAAAAAAAAAAEgAAACAEAAAAAAAIAAAACIAAAAEAAAQvAAQwC4AQAAAwAcAgKMCwAAAAAIAEAAJAAAAAAAAAgAoEAAAgAAAAAAgIAAAIAAAEQCAAAAKAQAAIAYGAAAwAAAAAQBBAAEACGAAFCAACQgACAH0owIAAAQADAAAIAgIJxgGECqAMOAgAAAAAAAAAAAAAEAAgBAAAASgAAAIIABADwABYAAA0AxEAIAgwEAARQAQAAAAAAAAAAAQAIEAAEDABMARMAABAAAAAAAAABIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAoAAAAAAAAAAAAAAAAAAAAAAAg%2Fbr%3D1%2Frs%3DACT90oHfnb-KzVr1I5MuNui3tBfqQ-5yPA,_basecomb:%2Fxjs%2F_%2Fjs%2Fk%3Dxjs.s.zh.ZEDG-AvAx4s.2018.O%2Fck%3Dxjs.s.1RouNDlYloc.L.B1.O%2Fam%3DAKQOAEIAAAAIAACIABUABAAAAAAAAAAAAAAAAAAAAAAAAEgAAACAEAAAAAAAIAAACKKAgAAEAAAQvAIQwC4AQAAAwAcAgKMCwAAAQAICEAgJAAAQAAAAAgAsEACAwAACBAAgIgAAIAAIEQCAAAgaAez3PwcGAAAwAAAAAQBDACECCGAAFC4ACQgADQH0owIIAAQADAAAIAgIJxgGECqAMPBgAAAAACAAAAAAAEAAgBAAAASg9AMIIABADwABYAAA0AxEAIAgwGgARQAS-AEACAAAAAAQAIEAIGDABMARMAABEAAAAAAA4D4AeDwgHFJYAAAAAAAAAAAAAAAAQAASBHNA-gsCQAAAAAAAAAAAAAAAAAAAQIqgicsNAAo%2Fd%3D1%2Fed%3D1%2Fdg%3D0%2Fbr%3D1%2Fujg%3D1%2Frs%3DACT90oGy8hrBykHbgzPn2HktrSfUNp1pzw'

def get_google_img(keyword):
    # url = 'https://www.google.com.hk/search?q={}&newwindow=1&sca_esv=280a44b9c274f648&sxsrf=ADLYWILBZhJl-s-ZfhbhYy1BxRLm1z1l2g:1729215146667&source=hp&ei=qroRZ4ThJvKm2roPwJwm&iflsig=AL9hbdgAAAAAZxHIurtuUS9h6TGCuPmKzuuYqq-Go1qy&udm=2'.format(keyword)
    url = 'https://www.google.com.hk/search?sca_esv=280a44b9c274f648&q={}&udm=2&fbs=AEQNm0AaBOazvTRM_Uafu9eNJJzCjPEAP5HX2BE31zy5nlFpWvns3PnbKLxkXEhI71m1WxuyGXhNUouYPTHnpIch5xJgIMcZof5GjjJE9giO3Q868M8bOV1cSsJN7w0Zrg98RFguyHOZm3xFt2oGVWl-pJTEpoK9LGg_7usrFKfWPNc2HJFGR3WY8McCLmP6UHbReplFnP4m&sa=X&ved=2ahUKEwje3qSzn-WJAxXwZfUHHc5oIEAQtKgLegQIEhAB&biw=1364&bih=753&dpr=2'.format(keyword)
    response = requests.get(url, headers=headers)
    bs = BeautifulSoup(response.text, 'html5lib')
    # print('213124:', response.text)
    # 获取top3的图片
    top3_ids = []
    top3_imgs = []
    top3_imgs2 = []
    print(len(bs.select('#search g-img')))
    for img in bs.select('#search g-img'):
        print(img.select('img')[0])
        img_id = img.select('img')[0].get('id')
        if len(img.select('img')[0].get('width')) >= 3:
            top3_ids.append(img_id)
    # 获取完整的base64 img
    for pic_id in top3_ids:
        for i in bs.select('script'):
            if 'data:image' in str(i) and "var ii=['{}']".format(pic_id) in str(i):
                base64_img = str(i).split('var s=')[1].split("';var")[0].replace("'", '')
                top3_imgs2.append(base64_img)
                if base64_img not in top3_imgs:
                    top3_imgs.append(base64_img)
    print('999999999:', len(top3_imgs), len(top3_imgs2))
    return top3_imgs

def save_comments_csv(data_arr):
    with open('comments_new_6.csv', 'a+', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerow(data_arr)  # 向文件中追加字符串列表
    print('insert ok！！！')

# 创建空目录
def create_path(your_path, base64_image_lists):
    # 假设固定目录为 images 文件夹，位于当前工作目录下
    fixed_directory = './图片/{}'.format(your_path)
    # 如果目录不存在则创建
    if not os.path.exists(fixed_directory):
        os.makedirs(fixed_directory)
    # Base64 编码的图片数据（这里只是示例数据，实际应用中你需要从其他地方获取这个数据）
    for base64_image_data in base64_image_lists:
        base64_image_data = base64_image_data.replace('\\x3d', '=')
        # 提取图像格式
        image_format = base64_image_data.split(';')[0].split('/')[1]
        # 解码 Base64 数据
        image_data = base64.b64decode(base64_image_data.split(',')[1])
        global num
        num += 1
        # 生成唯一的文件名，这里简单使用当前时间戳
        import time
        timestamp = str(int(time.time()))
        file_name = f"{timestamp}_{num}.{image_format}"
        # 完整的文件路径
        file_path = os.path.join(fixed_directory, file_name)
        # 保存图片
        with open(file_path, "wb") as f:
            f.write(image_data)
        print(f"图片已保存到 {file_path}")

data_lists = read_xlsx()
all_lens = len(data_lists)
# 序号从1开始，不算表头
# for i in range(1, all_lens):
for i in range(8, 9):
    keyword1 = str(data_lists[i])
    print('232432:', keyword1)
    keyword = quote(keyword1.replace('&', ''))
    base64_image_lists = get_google_img(keyword)
    # print(base64_image_lists)
    # save_comments_csv([i, keyword1, base64_image_lists])
    create_path(keyword1, base64_image_lists)


